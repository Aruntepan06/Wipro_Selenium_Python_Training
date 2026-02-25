# import openpyxl
#
# def get_excel_data(file_path, sheet_name):
#     # open the work book
#     workbook = openpyxl.load_workbook(file_path)
#
#     # fetch the sheet name
#     sheet = workbook[sheet_name]
#
#     data = []
#
#     for row in range(2, sheet.max_row + 1):
#         username = sheet.cell(row=row, column=1).value
#         password = sheet.cell(row=row, column=2).value
#         data.append((username, password))
#
#     return data

import openpyxl

def get_excel_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []

    for row in range(2, sheet.max_row + 1):
        row_data = []

        for col in range(1, sheet.max_column + 1):
            row_data.append(sheet.cell(row=row, column=col).value)

        data.append(tuple(row_data))

    return data