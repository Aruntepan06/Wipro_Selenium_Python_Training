"""
excel_utils.py

Utility module used for reading test data from Excel files.

This helper function supports data-driven testing in the automation
framework by extracting values from an Excel sheet and returning them
in a format compatible with pytest parameterization.

Libraries Used:
- openpyxl : Used to read Excel (.xlsx) files

Functionality:
Reads Excel workbook
Accesses specified sheet
Extracts row data (excluding header)
Returns data as list of tuples for pytest.parametrize
"""

import openpyxl


def get_excel_data(file_path, sheet_name):
    """
    Reads test data from an Excel sheet.

    Parameters:
    file_path : str
        Absolute or relative path of the Excel file.

    sheet_name : str
        Name of the sheet from which data should be read.

    Returns:
    list of tuples
        Each tuple represents one row of test data.
        This format is compatible with pytest.mark.parametrize.
    """

    try:
        # Load Excel workbook
        workbook = openpyxl.load_workbook(file_path)

        # Check if sheet exists
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")

        # Access the specified sheet
        sheet = workbook[sheet_name]

        # List to store extracted data
        data = []

        # Iterate over rows starting from row 2 (row 1 assumed to be header)
        for row in range(2, sheet.max_row + 1):

            row_data = []

            # Iterate through all columns in the current row
            for col in range(1, sheet.max_column + 1):

                # Fetch cell value and append to row_data
                row_data.append(sheet.cell(row=row, column=col).value)

            # Convert row data to tuple and append to data list
            data.append(tuple(row_data))

        # Return list of tuples
        return data

    except FileNotFoundError:
        print(f"ERROR: Excel file not found at path -> {file_path}")
        raise

    except ValueError as ve:
        print(f"ERROR: {ve}")
        raise

    except Exception as e:
        print(f"Unexpected error while reading Excel file: {e}")
        raise